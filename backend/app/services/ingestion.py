"""
Ingestion service for Module 2: BYO Retrieval.
Pipeline: update status → extract text → chunk → embed (batch) → insert chunks → update status
"""
import csv
import io
import os
import json
import logging
import re
import tempfile
import time
from typing import List

from google import genai

from app.services import retrieval_flags
from app.services.record_manager import compute_chunk_hash, compute_file_hash

logger = logging.getLogger(__name__)
EMBEDDING_MODEL = "models/gemini-embedding-001"
# 1536, not 768 (migration 022). The old 768 cited an ivfflat limit that migration 021
# removed when it moved the index to HNSW. Not 3072 (the model's native size) because
# pgvector's HNSW index tops out at 2000 dims — above that every search degrades to a
# full sequential scan.
EMBEDDING_DIMS = 1536

# Gemini embeds a question and a document differently when told which is which.
# Without this, a query's nearest neighbour is often another QUESTION rather than the
# passage that answers it. Both sides must stay consistent — changing either one
# requires re-embedding the corpus.
TASK_TYPE_DOCUMENT = "RETRIEVAL_DOCUMENT"
TASK_TYPE_QUERY = "RETRIEVAL_QUERY"
_client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))


def _convert_pptx_to_pdf(pptx_path: str) -> str:
    """Convert PPTX to PDF using Microsoft PowerPoint COM automation.

    Returns the path to the generated PDF, or empty string on failure.
    PowerPoint must be installed on the system.
    """
    import pythoncom
    import win32com.client

    pdf_path = pptx_path.rsplit(".", 1)[0] + ".pdf"
    abs_pptx = os.path.abspath(pptx_path)
    abs_pdf = os.path.abspath(pdf_path)

    try:
        pythoncom.CoInitialize()
        powerpoint = win32com.client.Dispatch("PowerPoint.Application")

        presentation = powerpoint.Presentations.Open(
            abs_pptx, ReadOnly=True, WithWindow=False,
        )
        # SaveAs with formatType=32 = ppSaveAsPDF
        presentation.SaveAs(abs_pdf, 32)
        presentation.Close()
        powerpoint.Quit()

        logger.info(f"Converted PPTX to PDF: {abs_pdf}")
        return abs_pdf
    except Exception as e:
        logger.error(f"PowerPoint COM conversion failed: {e}")
        return ""
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def extract_text(file_content: bytes, mime_type: str, file_name: str) -> str:
    ext = os.path.splitext(file_name)[1].lower()

    # Plain text formats — no parsing needed
    if ext in (".txt", ".md", ".csv", ".xml"):
        return file_content.decode("utf-8", errors="replace")

    # JSON — pretty-print
    if ext == ".json" or mime_type == "application/json":
        try:
            return json.dumps(json.loads(file_content.decode("utf-8", errors="replace")), indent=2)
        except Exception:
            return file_content.decode("utf-8", errors="replace")

    # PPTX/PPT: convert to PDF via PowerPoint first, then process as PDF
    # This gives us: perfect rendering, OCR on images, and slide numbers = page numbers
    if ext in (".pptx", ".ppt"):
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name
        pdf_path = ""
        try:
            pdf_path = _convert_pptx_to_pdf(tmp_path)
            if pdf_path and os.path.exists(pdf_path):
                # Process the PDF through Docling with OCR
                from docling.document_converter import DocumentConverter, PdfFormatOption
                from docling.datamodel.base_models import InputFormat
                from docling.datamodel.pipeline_options import PdfPipelineOptions

                converter = DocumentConverter(
                    format_options={
                        InputFormat.PDF: PdfFormatOption(
                            pipeline_options=PdfPipelineOptions(do_ocr=True),
                        ),
                    }
                )
                result = converter.convert(pdf_path)
                text = result.document.export_to_markdown()
                if text.strip():
                    logger.info(f"Processed PPTX via PDF conversion: {file_name}")
                    return text
                logger.warning(f"PDF conversion produced empty text for {file_name}, falling back to direct PPTX parsing")
            else:
                logger.warning(f"PPTX→PDF conversion failed for {file_name}, falling back to direct PPTX parsing")
        except Exception as e:
            logger.warning(f"PPTX→PDF pipeline failed for {file_name}: {e}")
        finally:
            os.unlink(tmp_path)
            if pdf_path and os.path.exists(pdf_path):
                os.unlink(pdf_path)
        # Fall through to direct Docling parsing if PDF conversion failed

    # Rich formats (PDF, DOCX, XLSX, HTML, images, etc.) — use docling
    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_content)
        tmp_path = tmp.name
    try:
        # Enable OCR for PDFs (extracts text from scanned pages / images)
        converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_options=PdfPipelineOptions(do_ocr=True),
                ),
            }
        )
        result = converter.convert(tmp_path)
        text = result.document.export_to_markdown()
        if not text.strip():
            # Fallback to UTF-8 decode if docling returns nothing
            return file_content.decode("utf-8", errors="replace")
        return text
    except Exception as e:
        logger.warning(f"Docling conversion failed for {file_name}: {e}")
        # If docling fails, fall back to raw decode
        return file_content.decode("utf-8", errors="replace")
    finally:
        os.unlink(tmp_path)


def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
    words = text.split()
    if not words:
        return []
    chunks, start = [], 0
    while start < len(words):
        end = min(start + chunk_size, len(words))
        chunk = " ".join(words[start:end])
        if chunk.strip():
            chunks.append(chunk)
        if end == len(words):
            break
        start = end - overlap
    return chunks


def chunk_csv(text: str, chunk_size: int = 500) -> List[str]:
    """Chunk a CSV on ROW boundaries, repeating the header in every chunk.

    The word-based splitter cuts a CSV wherever the word count lands — mid-value,
    mid-quoted-field — and only the first chunk ever carries the column names. A
    retrieved chunk then reads as bare values with no idea which column they are
    in or which entity the row describes. Cutting on rows and repeating the header
    makes every chunk self-describing.

    Returns [] when the text does not parse as a CSV with at least a header and one
    data row, so the caller can fall back to chunk_text() unchanged.
    """
    try:
        rows = list(csv.reader(io.StringIO(text)))
    except (csv.Error, ValueError):
        return []
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if len(rows) < 2:
        return []

    buf = io.StringIO()
    csv.writer(buf, lineterminator="\n").writerow(rows[0])
    header = buf.getvalue().rstrip("\n")
    header_words = len(header.split())
    if header_words >= chunk_size:
        return []  # pathological header - leave it to the word splitter

    chunks: List[str] = []
    current: List[str] = []
    current_words = header_words

    for row in rows[1:]:
        buf = io.StringIO()
        csv.writer(buf, lineterminator="\n").writerow(row)
        line = buf.getvalue().rstrip("\n")
        n = len(line.split())
        if current and current_words + n > chunk_size:
            chunks.append(header + "\n" + "\n".join(current))
            current, current_words = [], header_words
        current.append(line)
        current_words += n

    if current:
        chunks.append(header + "\n" + "\n".join(current))
    return chunks


def chunk_document(text: str, file_name: str = "", chunk_size: int = 500,
                   overlap: int = 50) -> List[str]:
    """Pick the chunker that suits the file, falling back to the word splitter."""
    if os.path.splitext(file_name or "")[1].lower() == ".csv":
        rows = chunk_csv(text, chunk_size=chunk_size)
        if rows:
            return rows
    return chunk_text(text, chunk_size=chunk_size, overlap=overlap)


IDENTITY_HEADER_MAX_LEN = 120
_PAGE_COLUMN_NAMES = {"source_page", "page", "page_no", "page_number"}


def identity_header(
    doc_title: str | None = None,
    file_name: str | None = None,
    section: str | None = None,
    page_start: int | None = None,
    page_end: int | None = None,
    tags: list[str] | None = None,
) -> str:
    """Build the `[...]` identity line prepended to a chunk's content when
    retrieval_flags.flag("chunk_identity") is on.

    Shape: `[<doc_title> · <file_name> · <section> · p<page> · tags: ...]`.
    Pure — same inputs always produce the same string; no clock, no
    randomness, no I/O. A field that is None/empty is left out entirely
    rather than leaving a bare ` · ` separator behind.
    """
    parts: list[str] = []
    if doc_title:
        parts.append(str(doc_title))
    if file_name:
        parts.append(str(file_name))
    if section:
        parts.append(str(section))

    if page_start is not None and page_end is not None and page_start != page_end:
        parts.append(f"p{page_start}-{page_end}")
    elif page_start is not None:
        parts.append(f"p{page_start}")
    elif page_end is not None:
        parts.append(f"p{page_end}")

    if tags:
        parts.append("tags: " + ", ".join(str(t) for t in tags))

    header = "[" + " · ".join(parts) + "]"
    if len(header) > IDENTITY_HEADER_MAX_LEN:
        # Defensive cap only — realistic corpus headers land near 100 chars
        # and never hit this. Keeps the ≤120-char contract even for a
        # pathological file_name/tags combination.
        header = header[: IDENTITY_HEADER_MAX_LEN - 1].rstrip(" ·") + "]"
    return header


def _humanize(segment: str) -> str:
    """'hwu_om_firefighting' -> 'Hwu Om Firefighting'. Pure string transform,
    no project-specific knowledge — used to turn a path/file segment into a
    human-readable label for the identity header."""
    words = re.sub(r"[_\-]+", " ", segment).split()
    return " ".join(w.capitalize() for w in words)


def _derive_doc_title(folder_path: str | None, file_name: str) -> str | None:
    """One folder = one logical document (doc-prep convention: 'One folder
    per source document'), so the folder name is the stable identity shared
    by every chunk of every file inside it — unlike file_name, which only
    identifies the one CSV/markdown file within that document. Falls back to
    the file's own name when there is no folder (root, or folder_path
    unknown).

    Deliberately NOT derived from LLM-extracted metadata (e.g. the 'topic'
    field): that call can return different wording on different ingests of
    the same file, which would make the embedded identity string — and
    therefore the embedding itself — silently non-reproducible run to run.
    """
    if folder_path and folder_path.strip("/"):
        segment = folder_path.rstrip("/").rsplit("/", 1)[-1]
        if segment:
            return _humanize(segment)
    stem = os.path.splitext(file_name or "")[0]
    return _humanize(stem) if stem else None


def _csv_chunk_page_range(chunk_text: str) -> tuple[int | None, int | None]:
    """Best-effort (min, max) page numbers for a CSV row-chunk, read from a
    'source_page'-shaped column repeated in the chunk's own header row
    (chunk_csv() repeats the header in every chunk — see its docstring).
    Pure — parses only the given string, no filesystem/network access.
    Returns (None, None) when the chunk doesn't parse as a header+rows CSV or
    carries no recognisable page column, so callers can omit the page
    segment of the identity header cleanly instead of guessing."""
    try:
        rows = list(csv.reader(io.StringIO(chunk_text)))
    except (csv.Error, ValueError):
        return (None, None)
    if len(rows) < 2:
        return (None, None)

    header = rows[0]
    page_idx = None
    for i, col in enumerate(header):
        if (col or "").strip().lower() in _PAGE_COLUMN_NAMES:
            page_idx = i
            break
    if page_idx is None:
        return (None, None)

    pages = []
    for row in rows[1:]:
        if page_idx < len(row):
            val = (row[page_idx] or "").strip()
            if val.isdigit():
                pages.append(int(val))
    if not pages:
        return (None, None)
    return (min(pages), max(pages))


def _is_rate_limit_error(e: Exception) -> bool:
    """Check if an exception is a 429 rate limit error."""
    err_str = str(e)
    return "429" in err_str or "RESOURCE_EXHAUSTED" in err_str


def _embed_with_retry(func, *args, max_retries: int = 5, **kwargs):
    """Call an embedding function with exponential backoff on failure."""
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            if attempt == max_retries - 1:
                raise
            if _is_rate_limit_error(e):
                wait = 5 * (2 ** attempt)  # 5, 10, 20, 40, 80s for rate limits
            else:
                wait = 2 ** (attempt + 1)  # 2, 4, 8, 16, 32s for other errors
            logger.warning(f"Embedding attempt {attempt + 1} failed: {e}. Retrying in {wait}s...")
            time.sleep(wait)


def embed_text(text: str, task_type: str = TASK_TYPE_DOCUMENT, title: str | None = None) -> List[float]:
    """Embed one string. Defaults to DOCUMENT — callers embedding a user's question
    must pass TASK_TYPE_QUERY.

    `title` is Gemini's per-document identity hint (gemini-embedding-001, docs say
    it improves retrieval quality). It is only valid for RETRIEVAL_DOCUMENT, so it
    is silently dropped for any other task_type instead of being sent and risking
    an API-side rejection or no-op. Omitted (None) reproduces today's request
    exactly — this is what keeps chunk_identity's flag-off path byte-identical."""
    config = {"output_dimensionality": EMBEDDING_DIMS, "task_type": task_type}
    if title and task_type == TASK_TYPE_DOCUMENT:
        config["title"] = title
    response = _embed_with_retry(
        _client.models.embed_content,
        model=EMBEDDING_MODEL, contents=text,
        config=config,
    )
    return response.embeddings[0].values


def _split_by_token_budget(texts: List[str], max_tokens: int = 18000, max_items: int = 50) -> List[List[str]]:
    """Split texts into batches that stay under the Gemini embedding token limit.
    Uses ~4 chars per token as a conservative estimate."""
    batches = []
    current_batch = []
    current_tokens = 0
    for text in texts:
        est_tokens = len(text) // 4 + 1
        if current_batch and (current_tokens + est_tokens > max_tokens or len(current_batch) >= max_items):
            batches.append(current_batch)
            current_batch = []
            current_tokens = 0
        current_batch.append(text)
        current_tokens += est_tokens
    if current_batch:
        batches.append(current_batch)
    return batches


def embed_batch(texts: List[str], batch_size: int = 50, title: str | None = None) -> List[List[float]]:
    """Embed a batch of DOCUMENT-side texts. `title` (optional) is passed to Gemini
    once per batch call, which is correct as long as every text in one embed_batch()
    call belongs to the same source document — true for every current caller
    (ingest_document / ingest_document_update process one file at a time).
    Omitted by default, so existing callers (reembed_all.py etc.) are unaffected."""
    if not texts:
        return []
    all_embeddings = []
    batches = _split_by_token_budget(texts, max_tokens=18000, max_items=batch_size)
    for idx, batch in enumerate(batches):
        try:
            config = {"output_dimensionality": EMBEDDING_DIMS, "task_type": TASK_TYPE_DOCUMENT}
            if title:
                config["title"] = title
            response = _embed_with_retry(
                _client.models.embed_content,
                model=EMBEDDING_MODEL, contents=batch,
                config=config,
            )
            all_embeddings.extend([e.values for e in response.embeddings])
        except Exception:
            # Fallback: embed one at a time for this batch
            all_embeddings.extend([embed_text(t, TASK_TYPE_DOCUMENT, title=title) for t in batch])
        # Rate-limit pause between batches to avoid 429 errors
        if idx < len(batches) - 1:
            time.sleep(1)
    return all_embeddings


MAX_ROWS_PER_SHEET = 10_000
MAX_HEADER_SCAN_ROWS = 10  # How many rows to scan looking for headers


def _sanitize_table_name(file_name: str, sheet_title: str = "") -> str:
    base = os.path.splitext(file_name)[0]
    name = f"{base}_{sheet_title}" if sheet_title else base
    name = re.sub(r"[^a-zA-Z0-9_]", "_", name.lower()).strip("_")
    return re.sub(r"_+", "_", name)


def _score_header_row(row: tuple, total_cols: int) -> float:
    """Score how likely a row is to be a header. Higher = more likely."""
    if not row:
        return -1
    non_null = [v for v in row if v is not None]
    if not non_null:
        return -1

    score = 0.0
    # Headers should have many non-null values relative to total columns
    fill_ratio = len(non_null) / max(total_cols, 1)
    score += fill_ratio * 40  # 0-40 points

    text_count = 0
    numeric_count = 0
    short_count = 0
    for v in non_null:
        s = str(v).strip()
        if not s:
            continue
        # Headers are typically short text labels
        if isinstance(v, str) or (not isinstance(v, (int, float))):
            text_count += 1
        else:
            numeric_count += 1
        if len(s) <= 30:
            short_count += 1

    if non_null:
        # Headers should be mostly text, not numbers
        score += (text_count / len(non_null)) * 30  # 0-30 points
        # Headers should be short strings
        score += (short_count / len(non_null)) * 20  # 0-20 points
        # Bonus: many distinct values (not all the same)
        unique_ratio = len(set(str(v) for v in non_null)) / len(non_null)
        score += unique_ratio * 10  # 0-10 points

    return score


def _extract_structured_data(
    file_content: bytes, file_name: str, document_id: str, user_id: str, supabase_client
) -> None:
    """Extract tabular data from CSV/XLSX and store in structured_data table."""
    ext = os.path.splitext(file_name)[1].lower()

    tables = []  # list of (table_name, columns, rows)

    if ext == ".csv":
        text = file_content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        columns = reader.fieldnames or []
        if not columns:
            return
        rows = []
        for i, row in enumerate(reader):
            if i >= MAX_ROWS_PER_SHEET:
                break
            rows.append(dict(row))
        if rows:
            tables.append((_sanitize_table_name(file_name), list(columns), rows))

    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        for sheet in wb.worksheets:
            sheet_rows = list(sheet.iter_rows(values_only=True))
            if not sheet_rows or len(sheet_rows) < 2:
                continue

            # Smart header detection: scan first N rows, pick the best header
            total_cols = len(sheet_rows[0])
            scan_limit = min(MAX_HEADER_SCAN_ROWS, len(sheet_rows) - 1)  # need at least 1 data row after
            best_idx = 0
            best_score = -1
            for idx in range(scan_limit):
                score = _score_header_row(sheet_rows[idx], total_cols)
                if score > best_score:
                    best_score = score
                    best_idx = idx

            header_row_idx = best_idx
            data_start_idx = header_row_idx + 1
            if header_row_idx > 0:
                logger.info(f"Sheet '{sheet.title}': detected header at row {header_row_idx} (score={best_score:.1f})")

            raw_headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(sheet_rows[header_row_idx])]
            if not any(h.strip() for h in raw_headers if h):
                continue

            # Sanitize header names: lowercase, replace special chars, deduplicate
            seen = {}
            clean_headers = []
            for h in raw_headers:
                sanitized = re.sub(r"[^a-zA-Z0-9_]", "_", h).strip("_")
                sanitized = re.sub(r"_+", "_", sanitized).lower()
                if not sanitized or sanitized.startswith("col_"):
                    sanitized = f"col_{len(clean_headers)}"
                # Handle duplicate column names
                if sanitized in seen:
                    seen[sanitized] += 1
                    sanitized = f"{sanitized}_{seen[sanitized]}"
                else:
                    seen[sanitized] = 0
                clean_headers.append(sanitized)

            data_rows = []
            for row in sheet_rows[data_start_idx:MAX_ROWS_PER_SHEET + data_start_idx]:
                data_rows.append({h: v for h, v in zip(clean_headers, row)})
            if data_rows:
                tables.append((_sanitize_table_name(file_name, sheet.title), clean_headers, data_rows))
        wb.close()

    for table_name, columns, rows in tables:
        # JSON-serialize values that aren't natively JSON-compatible
        clean_rows = []
        for row in rows:
            clean = {}
            for k, v in row.items():
                if v is None:
                    clean[k] = None
                elif isinstance(v, (int, float, bool, str)):
                    clean[k] = v
                else:
                    clean[k] = str(v)
            clean_rows.append(clean)

        supabase_client.table("structured_data").insert({
            "document_id": document_id,
            "user_id": user_id,
            "table_name": table_name,
            "columns": columns,
            "rows": clean_rows,
            "row_count": len(clean_rows),
        }).execute()
        logger.info(f"Stored structured data '{table_name}' ({len(clean_rows)} rows) for document {document_id}")


def _embed_and_build_chunk_rows(
    chunks: List[str], document_id: str, user_id: str, file_name: str, folder_path: str | None,
) -> List[dict]:
    """Embed `chunks` and build the document_chunks row dicts to insert.

    Shared by ingest_document and ingest_document_update so the two ingest
    paths cannot diverge (both must apply chunk_identity identically).

    Flag OFF: reproduces pre-chunk_identity behaviour byte-for-byte — same
    `content`, same embed_content config (no `title`), no identity columns
    populated. This is the guard the flag-off test asserts.
    """
    if not retrieval_flags.flag("chunk_identity"):
        embeddings = embed_batch(chunks)
        return [
            {
                "document_id": document_id,
                "user_id": user_id,
                "content": chunk,
                "embedding": embedding,
                "chunk_index": idx,
                "content_hash": compute_chunk_hash(chunk),
            }
            for idx, (chunk, embedding) in enumerate(zip(chunks, embeddings))
        ]

    doc_title = _derive_doc_title(folder_path, file_name)
    stamped_chunks: List[str] = []
    page_ranges: List[tuple] = []
    for chunk in chunks:
        page_start, page_end = _csv_chunk_page_range(chunk)
        header = identity_header(
            doc_title=doc_title, file_name=file_name, section=None,
            page_start=page_start, page_end=page_end, tags=None,
        )
        stamped_chunks.append(f"{header}\n{chunk}")
        page_ranges.append((page_start, page_end))

    embeddings = embed_batch(stamped_chunks, title=doc_title)

    return [
        {
            "document_id": document_id,
            "user_id": user_id,
            "content": stamped,
            "embedding": embedding,
            "chunk_index": idx,
            "content_hash": compute_chunk_hash(stamped),
            "file_name": file_name,
            "folder_path": folder_path,
            "doc_title": doc_title,
            "page_start": page_ranges[idx][0],
            "page_end": page_ranges[idx][1],
            "section_path": None,
            "tags": None,
        }
        for idx, (stamped, embedding) in enumerate(zip(stamped_chunks, embeddings))
    ]


def _folder_path_for(document_id: str, supabase_client) -> str | None:
    """One extra SELECT, made only when chunk_identity is on — folder_path is
    not part of ingest_document's own parameters (both call sites insert the
    documents row, with folder_path, before calling ingest_document)."""
    row = supabase_client.table("documents").select("folder_path").eq(
        "id", document_id).single().execute().data or {}
    return row.get("folder_path")


def ingest_document(
    document_id: str,
    file_content: bytes,
    mime_type: str,
    file_name: str,
    user_id: str,
    supabase_client,
) -> None:
    try:
        supabase_client.table("documents").update(
            {"status": "processing", "updated_at": "now()"}
        ).eq("id", document_id).execute()

        text = extract_text(file_content, mime_type, file_name)
        if not text.strip():
            raise ValueError("No extractable text found in document")

        # Extract metadata before chunking
        try:
            from app.services.metadata import extract_metadata
            metadata = extract_metadata(text, file_name=file_name, file_content=file_content)
            supabase_client.table("documents").update(
                {"metadata": metadata, "updated_at": "now()"}
            ).eq("id", document_id).execute()
            logger.info(f"Extracted metadata for document {document_id}")
        except Exception as e:
            logger.warning(f"Metadata extraction failed (non-fatal) for {document_id}: {e}")

        # Extract structured data for CSV/XLSX (non-fatal)
        try:
            _extract_structured_data(file_content, file_name, document_id, user_id, supabase_client)
        except Exception as e:
            logger.warning(f"Structured data extraction failed (non-fatal) for {document_id}: {e}")

        chunks = chunk_document(text, file_name, chunk_size=500, overlap=50)
        if not chunks:
            raise ValueError("Chunking produced no chunks")

        folder_path = (
            _folder_path_for(document_id, supabase_client)
            if retrieval_flags.flag("chunk_identity") else None
        )
        rows = _embed_and_build_chunk_rows(chunks, document_id, user_id, file_name, folder_path)
        for i in range(0, len(rows), 100):
            supabase_client.table("document_chunks").insert(rows[i : i + 100]).execute()

        file_hash = compute_file_hash(file_content)
        supabase_client.table("documents").update({
            "status": "ready",
            "content_hash": file_hash,
            "content_markdown": text,                  # BACKFILL-01: synchronous markdown capture
            "content_markdown_status": "ready",        # BACKFILL-01: same UPDATE = atomic
            "updated_at": "now()",
        }).eq("id", document_id).execute()
        logger.info(
            f"Ingested document {document_id}: {len(chunks)} chunks, "
            f"{len(text)} markdown chars"
        )

    except Exception as e:
        error_msg = str(e)
        logger.error(f"Ingestion failed for document {document_id}: {error_msg}")
        try:
            supabase_client.table("documents").update({
                "status": "failed",
                "content_markdown_status": "failed",   # BACKFILL-04: surface to Phase 4 tools
                "error_message": error_msg,
                "updated_at": "now()",
            }).eq("id", document_id).execute()
        except Exception as inner_e:
            logger.error(f"Could not update failed status: {inner_e}")


def ingest_document_update(
    document_id: str,
    file_content: bytes,
    mime_type: str,
    file_name: str,
    user_id: str,
    supabase_client,
) -> None:
    """Re-ingest a document: delete all old chunks, re-chunk and re-embed from scratch."""
    try:
        supabase_client.table("documents").update(
            {"status": "processing", "updated_at": "now()"}
        ).eq("id", document_id).execute()

        # Delete all existing chunks and structured data
        supabase_client.table("document_chunks").delete().eq("document_id", document_id).execute()
        supabase_client.table("structured_data").delete().eq("document_id", document_id).execute()

        text = extract_text(file_content, mime_type, file_name)
        if not text.strip():
            raise ValueError("No extractable text found in document")

        # Re-extract metadata on update
        try:
            from app.services.metadata import extract_metadata
            metadata = extract_metadata(text, file_name=file_name, file_content=file_content)
            supabase_client.table("documents").update(
                {"metadata": metadata, "updated_at": "now()"}
            ).eq("id", document_id).execute()
            logger.info(f"Re-extracted metadata for document {document_id}")
        except Exception as e:
            logger.warning(f"Metadata extraction failed (non-fatal) for {document_id}: {e}")

        # Re-extract structured data for CSV/XLSX (non-fatal)
        try:
            _extract_structured_data(file_content, file_name, document_id, user_id, supabase_client)
        except Exception as e:
            logger.warning(f"Structured data extraction failed (non-fatal) for {document_id}: {e}")

        chunks = chunk_document(text, file_name, chunk_size=500, overlap=50)
        if not chunks:
            raise ValueError("Chunking produced no chunks")

        folder_path = (
            _folder_path_for(document_id, supabase_client)
            if retrieval_flags.flag("chunk_identity") else None
        )
        rows = _embed_and_build_chunk_rows(chunks, document_id, user_id, file_name, folder_path)
        for i in range(0, len(rows), 100):
            supabase_client.table("document_chunks").insert(rows[i:i+100]).execute()

        file_hash = compute_file_hash(file_content)
        supabase_client.table("documents").update({
            "status": "ready",
            "content_hash": file_hash,
            "content_markdown": text,                  # BACKFILL-01: synchronous markdown capture (re-ingest path)
            "content_markdown_status": "ready",        # BACKFILL-01: same UPDATE = atomic
            "updated_at": "now()",
        }).eq("id", document_id).execute()

        logger.info(
            f"Re-ingested document {document_id}: {len(chunks)} chunks, "
            f"{len(text)} markdown chars"
        )
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Update ingestion failed for {document_id}: {error_msg}")
        try:
            supabase_client.table("documents").update({
                "status": "failed",
                "content_markdown_status": "failed",   # BACKFILL-04: surface to Phase 4 tools
                "error_message": error_msg,
                "updated_at": "now()",
            }).eq("id", document_id).execute()
        except Exception as inner_e:
            logger.error(f"Could not update failed status: {inner_e}")
